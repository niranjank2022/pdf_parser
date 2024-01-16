import collections
import os
import re
import openpyxl
from pdfminer.high_level import extract_text
from pathlib import Path
import pathlib
from playsound import playsound


def get_path():
    print("Hello there! Help me out with parsing your files.\n"
          "Kindly give me the path address of the folder with all the pdf files.")

    input_path = input("Path: ")
    while not Path(input_path).is_dir():
        print("Path doesn't exist. Try again!")
        input_path = input("Path: ")

    output_path = input("\nWhere do you want to store the output file?\nPath: ")
    while not Path(output_path).is_dir():
        print("Path doesn't exist. Try again!")
        output_path = input("Path: ")

    return input_path, output_path


def extract_data(input_path, output_path):
    path = Path(input_path)
    files = sorted(path.glob("*.pdf"))
    wb = openpyxl.load_workbook(output_path)
    sheet = wb.active
    completed = set()
    for row in sheet.iter_rows(values_only=True):
        completed.add(row[12])

    for file in files:
        if file.name in completed:
            continue
        print(file.name)
        # try:
        text = extract_text(file)
        data = extract_data_from_text(text)
        data.append(file.name)
        sheet.append(data)
        # except:
        #     print("Unable to parse " + file.name)

        wb.save(output_path)

    # playsound(r"bin\alert.wav")


def refactor_id(text, data):
    if 'O' in data['q-id'] or '0' in data['q-id']:
        data['q-id'] = data['q-id'].replace('O', '0')
        combinations = generate_combo(data['q-id'])
        for seq in combinations:
            if seq != data['q-id']:
                text = text.replace(seq, data['q-id'])

    return text


def generate_combo(txt):
    indices = []
    i = 0
    while i < len(txt) and '0' in txt[i:]:
        i = txt.index('0', i)
        indices.append(i)
        i += 1

    comb = [txt]
    for i in range(len(indices)):
        for j in range(len(comb)):
            e = comb.pop()
            comb.append(e[:indices[i]] + '0' + e[indices[i] + 1:])
            comb.append(e[:indices[i]] + 'O' + e[indices[i] + 1:])

    return comb


def extract_data_from_text(text):
    data = dict()
    text = remove_latin_ligatures(text)
    lines = text.strip().split("\n\n")
    text = "\n".join(lines)

    data['q-id'] = lines[0].split()[-1]
    text = refactor_id(text, data)
    re_text = text.replace("\n", "###")
    print(text)
    print(re_text)

    data['domain'] = lines[8].replace("\n", " ")
    data['skill'] = lines[9].replace("\n", " ")

    data['difficulty'] = lines[-1].split()[-1]
    q = re.findall(f"ID:(.*?)###A\\.", re_text)[0]
    q = q.split("###")[1:]

    data['prompt'] = q[-1].strip()
    data['question'] = "\n".join(q[:-1]).strip().replace("###", "\n").strip()
    data['choice-a'] = re.findall("###A\\.(.*?)###B\\.", re_text)[0].replace("###", "\n").strip()
    try:
        data['choice-b'] = re.findall("###B\\.(.*?)###C\\.", re_text)[0].replace("###", "\n").strip()
    except:
        data['choice-b'] = re.findall("###B\\.(.*?)### *C\\.", re_text)[0].replace("###", "\n").strip()
    try:
        data['choice-c'] = re.findall("###C\\.(.*?)###D\\.", re_text)[0].replace("###", "\n").strip()
    except:
        data['choice-c'] = re.findall("### *C\\.(.*?)###D\\.", re_text)[0].replace("###", "\n").strip()
    try:
        data['choice-d'] = re.findall("###D\\.(.*?)###ID:", re_text)[0].replace("###", "\n").strip()
    except:
        data['choice-d'] = re.findall("### *D\\.(.*?)###ID:", re_text)[0].replace("###", "\n").strip()

    data['answer'] = re.findall("Correct Answer: (.*?)###", re_text)[0].strip()
    try:
        data['rationale'] = re.findall("Rationale(.*?)" + lines[-1], re_text)[0].replace("###", "\n").strip()
    except:
        re_text += "&&&"
        data['rationale'] = re.findall("Rationale(.*?)&&&", re_text)[0].replace("###", "\n").strip()

    res = []
    for key in (
            'q-id', 'domain', 'skill', 'difficulty', 'question', 'prompt', 'choice-a', 'choice-b', 'choice-c',
            'choice-d',
            'answer', 'rationale'):
        if data[key] == '':
            print(key)
            raise ValueError
        res.append(data[key])

    return res


def remove_latin_ligatures(text):
    liga = {
        'ﬀ': 'ff', 'ﬁ': 'fi', 'ﬂ': 'fl', 'ﬃ': 'ffi', 'ﬄ': 'ffl', 'ﬅ': 'ft', 'ﬆ': 'st', 'Æ': 'AE', 'æ': 'ae', 'Ĳ': 'IJ',
        'ĳ': 'ij', 'Œ': 'OE', 'œ': 'oe'
    }
    s = ""
    for ch in text:
        if ord(ch) == 0 or ord(ch) == 160 or ord(ch) == 12:
            s += ' '
            continue
        s += liga.get(ch, ch)

    return s


def segregate_failed_files(xlsx_path):
    wb = openpyxl.load_workbook(output_path)
    sheet = wb.active
    failed = []
    for row in sheet.iter_rows(values_only=True):
        for val in row:
            if val is None:
                failed.append(row[-1])
                path = pathlib.Path(row[-1]).absolute()
                if path.exists():
                    parts = list(path.parts[:-1]) + ['failed', row[-1]]
                    new_path = Path(*parts)
                    parts[-2] = 'file_dump'
                    path = Path(*parts)
                    os.rename(path, new_path)
                break
    return failed


if __name__ == '__main__':
    input_path = "file_dump/"
    output_path = "output_excel/CB SAT RW Question Bank.xlsx"
    extract_data(input_path, output_path)
    # segregate_failed_files(output_path)
